from openpyxl import load_workbook, Workbook

wb = load_workbook('/home/knight/books.xlsx')

# Get a sheet by name
# sheet = wb.get_sheet_by_name('Sheet3')
sheet = wb['Sheet1']
wb2 = load_workbook('/home/knight/books.xlsx')

print(wb2.sheetnames)

sheet_ranges = wb['Sheet1']
print(sheet_ranges['d1'].value)
i = 0
print(sheet_ranges.cell(column='a1', row='a1'))
# while i < 10:
#     print(i)
#     z='a'+str(i)
#     i+=1
#     print(sheet_ranges['a3'].value)

